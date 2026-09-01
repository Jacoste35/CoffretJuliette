#!/usr/bin/env python3
"""Extraction FIDELE du classeur fournisseur 'Tarif et bon de commande 2026'.

Ce script ne transforme rien : il recopie les cellules telles quelles dans
deux CSV bruts, en propageant les notes fusionnees (colonne J) sur toutes les
lignes de leur plage. Il sert a tracer l'origine exacte de chaque donnee.

Usage : python3 extract_xlsx.py <chemin_du_xlsx> <dossier_raw>
"""
import csv
import sys
from pathlib import Path

import openpyxl


def note_map(ws, col_letter="J"):
    """Retourne {no_ligne: note} en propageant les cellules fusionnees."""
    notes = {}
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(row=rng.min_row, column=rng.min_col)
        if top_left.value is None:
            continue
        if openpyxl.utils.get_column_letter(rng.min_col) != col_letter:
            continue
        text = " / ".join(s.strip() for s in str(top_left.value).splitlines() if s.strip())
        for r in range(rng.min_row, rng.max_row + 1):
            notes[r] = text
    for row in ws.iter_rows():
        cell = ws[f"{col_letter}{row[0].row}"]
        if cell.value is not None and row[0].row not in notes:
            notes[row[0].row] = " / ".join(
                s.strip() for s in str(cell.value).splitlines() if s.strip()
            )
    return notes


def cell(ws, row, col):
    v = ws[f"{col}{row}"].value
    if v is None:
        return ""
    return str(v).strip()


def extract_normand_direct(ws, out_path):
    notes = note_map(ws)
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "ligne_source",
                "fournisseur_colonne_A",
                "dlc_ddm",
                "produit",
                "pu_ht",
                "multiple_de_commande",
                "tva",
                "note_fournisseur",
            ]
        )
        for row in range(10, ws.max_row + 1):
            produit = cell(ws, row, "C")
            fourn = cell(ws, row, "A")
            if not produit and not fourn:
                continue
            w.writerow(
                [
                    row,
                    fourn,
                    cell(ws, row, "B"),
                    produit,
                    cell(ws, row, "D"),
                    cell(ws, row, "E"),
                    cell(ws, row, "I"),
                    notes.get(row, ""),
                ]
            )


def extract_isigny(ws, out_path):
    """La feuille Isigny est une suite de blocs : titre de gamme, ligne
    d'en-tetes, puis articles. On conserve le titre de gamme courant et le
    sous-titre courant (ligne sans reference) pour chaque article."""
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "ligne_source",
                "gamme",
                "sous_gamme",
                "article",
                "reference",
                "poids",
                "prix_ht_colonne_G",
                "libelle_colonne_G",
                "tva",
                "ean",
                "colisage",
                "prix_colonne_K",
                "libelle_colonne_K",
            ]
        )
        gamme = sous_gamme = ""
        libelle_g = libelle_k = ""
        for row in range(1, ws.max_row + 1):
            b = cell(ws, row, "B")
            ref = cell(ws, row, "E")
            if not b and not ref:
                continue
            if b == "Article":
                # ligne d'en-tetes : on memorise le libelle exact des colonnes
                libelle_g = cell(ws, row, "G")
                libelle_k = cell(ws, row, "K")
                continue
            if not ref:
                # ligne de titre : gamme (contient "page") ou sous-gamme
                if "page" in b.lower():
                    gamme, sous_gamme = b, ""
                else:
                    sous_gamme = b
                continue
            # colonne TVA : H sur la plupart des blocs, I sur les blocs vrac
            tva = cell(ws, row, "H") or cell(ws, row, "I")
            ean = cell(ws, row, "I") if cell(ws, row, "H") else ""
            w.writerow(
                [
                    row,
                    gamme,
                    sous_gamme,
                    b,
                    ref,
                    cell(ws, row, "F"),
                    cell(ws, row, "G"),
                    libelle_g,
                    tva,
                    ean,
                    cell(ws, row, "K") if cell(ws, row, "H") else "",
                    cell(ws, row, "K") if not cell(ws, row, "H") else "",
                    libelle_k,
                ]
            )


def main():
    src, out_dir = Path(sys.argv[1]), Path(sys.argv[2])
    wb = openpyxl.load_workbook(src, data_only=True)
    extract_normand_direct(
        wb["TARIF NORMAND DIRECT 2026"], out_dir / "01_normand_direct_tarif_2026.csv"
    )
    extract_isigny(wb["CARAMELS ISIGNY 2026"], out_dir / "02_caramels_isigny_2026.csv")
    print("OK ->", out_dir)


if __name__ == "__main__":
    main()
