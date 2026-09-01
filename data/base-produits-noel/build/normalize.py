#!/usr/bin/env python3
"""Construit la base normalisee PRODUITS / FOURNISSEURS / TARIFS / ... a
partir des extractions brutes du dossier raw/.

Regle absolue du projet : aucune donnee inventee. Toute information absente
des documents sources vaut A CONFIRMER, et chaque ligne porte la source
exacte dont elle provient.
"""
import csv
import re
import sys
from collections import defaultdict, Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import rules as R  # noqa: E402
import pricing as PRICING  # noqa: E402

BASE = Path(__file__).resolve().parent.parent
RAW = BASE / "raw"
EXPORT = BASE / "export"
A = R.A_CONFIRMER

SRC_NORMAND = "Tarif et bon de commande 2026 – onglet « TARIF NORMAND DIRECT 2026 »"
SRC_ISIGNY = "Tarif et bon de commande 2026 – onglet « CARAMELS ISIGNY 2026 »"
SRC_BILLY = "billy.pdf – tarif SARL Domaine de Billy « Prix de vente à partir de 2026 »"
SRC_BRUCAN = "FACT_2600000025_THE_ART_OF_WINE_ACADEMY.pdf – facture Domaine de Brucan du 12/08/2026"


def read_csv(name):
    with open(RAW / name, encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def num(value):
    """Extrait le premier nombre du texte d'un document et le convertit en
    float. Les documents ecrivent les prix sous des formes tres variees :
    « 3,36 euros H.T. », « 35,50 € », « 45,50€/kg », « 2,30, € »."""
    if value in (None, ""):
        return ""
    s = str(value).replace("\u00a0", " ").replace("\u202f", " ").replace("€", " ")
    m = re.search(r"-?\d+(?:[ .]\d{3})*(?:[.,]\d+)?", s)
    if not m:
        return ""
    try:
        return round(float(m.group(0).replace(" ", "").replace(",", ".")), 4)
    except ValueError:
        return ""


# ==========================================================================
# 1. Fournisseurs
# ==========================================================================
_ref_rows = read_csv("../build/referentiel_fournisseurs.csv")
referentiel = {r["cle_source"]: r for r in _ref_rows}
# Index normalise : les documents melangent apostrophes droites et typographiques
# (« POM' POM' » et « POM’ POM’ » designent le meme producteur).
referentiel_normalise = {R.normalize_key(r["cle_source"]): r for r in _ref_rows}


def ref_producteur(nom):
    return referentiel.get(nom) or referentiel_normalise.get(R.normalize_key(nom))


DEPARTEMENTS = {
    "14": "Calvados (14)", "27": "Eure (27)", "50": "Manche (50)",
    "61": "Orne (61)", "76": "Seine-Maritime (76)",
}


def normaliser_departement(valeur):
    """Le referentiel melange « 14 » et « Calvados (14) »."""
    v = (valeur or "").strip().replace("Seine Maritime", "Seine-Maritime")
    return DEPARTEMENTS.get(v, v)

fournisseurs = {}   # nom_normalise -> dict
_next_four = [0]


def get_fournisseur(nom, role, canal, source, libelle_source=""):
    key = R.normalize_key(nom)
    if key in fournisseurs:
        f = fournisseurs[key]
        if libelle_source and libelle_source not in f["libelles_source"]:
            f["libelles_source"].append(libelle_source)
        return f
    ref = referentiel.get(nom) or referentiel_normalise.get(R.normalize_key(nom)) or {}
    _next_four[0] += 1
    f = {
        "id_fournisseur": f"FOUR-{_next_four[0]:04d}",
        "nom": nom,
        "libelles_source": [libelle_source] if libelle_source else [],
        "role": ref.get("role") or role,
        "canal_de_commande": canal,
        "regroupement_suggere": ref.get("regroupement_suggere") or R.regroupement(nom),
        "regroupement_statut": "À CONFIRMER (regroupement proposé, non appliqué)"
        if (ref.get("regroupement_suggere") or R.regroupement(nom))
        else "",
        "contact_nom": ref.get("contact_nom", A),
        "telephone": ref.get("telephone", A),
        "email": ref.get("email", A),
        "site_internet": ref.get("site_internet", A),
        "adresse": ref.get("adresse", A),
        "code_postal": ref.get("code_postal", A),
        "ville": ref.get("ville", A),
        "region": ref.get("region", A),
        "departement": ref.get("departement", A),
        "identifiants_legaux": ref.get("identifiants_legaux", A),
        "conditions_paiement": ref.get("conditions_paiement", A),
        "livraison": ref.get("livraison", A),
        "retrait": ref.get("retrait", A),
        "minimum_commande": ref.get("minimum_commande", A),
        "personnalisation": ref.get("personnalisation", A),
        "delais": ref.get("delais", A),
        "autres_infos": ref.get("autres_infos", ""),
        "conditions_relevees_dans_le_tarif": set(),
        "source": ref.get("source", source),
    }
    fournisseurs[key] = f
    return f


NORMAND = get_fournisseur(
    "NORMAND DIRECT TERROIR", "DISTRIBUTEUR / GROSSISTE",
    "Commande directe auprès de Normand Direct Terroir", SRC_NORMAND)
ISIGNY = get_fournisseur(
    "CARAMELS D'ISIGNY", "FABRICANT",
    "À CONFIRMER : onglet intégré au classeur Normand Direct Terroir, canal de commande non précisé",
    SRC_ISIGNY)
BILLY = get_fournisseur(
    "SARL DOMAINE DE BILLY", "PRODUCTEUR (vente directe)",
    "Commande directe auprès du producteur", SRC_BILLY)
BRUCAN = get_fournisseur(
    "DOMAINE DE BRUCAN", "PRODUCTEUR (vente directe)",
    "Commande directe auprès du producteur", SRC_BRUCAN)


# ==========================================================================
# 2. Produits
# ==========================================================================
produits = []
anomalies = []
_next_prod = [0]


def new_id():
    _next_prod[0] += 1
    return f"PROD-{_next_prod[0]:04d}"


def base_produit(**kw):
    """Fiche produit etendue. Tout champ non renseigne vaut A CONFIRMER."""
    p = {
        # identification
        "id_produit": "", "reference_fournisseur": A, "nom_produit": "",
        "libelle_source": "", "description": A, "categorie": A,
        "sous_categorie": A, "regle_de_classification": "", "marque": A,
        "id_producteur": "", "producteur": A, "producteur_libelle_source": "",
        "id_fournisseur": "", "fournisseur": "", "canal_de_commande": "",
        "url_photo": A,
        # origine
        "origine_pays": A, "origine_region": A, "origine_departement": A,
        "lieu_fabrication": A, "appellation": "", "niveau_preuve_origine": "",
        "caractere_local": A, "savoir_faire_francais": A,
        "alerte_matiere_premiere": "",
        # caracteristiques
        "poids_net_g": "", "volume_ml": "", "mention_poids_volume_source": "",
        "degre_alcool": "", "alcool": A, "type_alcool": "",
        "dlc_ddm": A, "mention_source": "",
        # prix d'achat
        "prix_achat_ht": "", "prix_achat_texte_source": "",
        "unite_de_prix": "unité", "libelle_colonne_prix_source": "",
        "tva_taux": A, "prix_achat_ttc": "", "prix_au_kg_ht": "",
        "droits_accises": "", "prix_hors_droits_ht": "",
        # prix de vente (a definir : aucun document ne les contient)
        "prix_vente_ht": "", "prix_vente_ttc": "", "marge_eur": "",
        "taux_marge": "", "prix_min_acceptable": "", "prix_max_conseille": "",
        "coefficient_applique": "", "calcul_prix_vente": "", "statut_prix_vente": "",
        # conditionnement / commande
        "conditionnement": A, "quantite_par_conditionnement": "",
        "multiple_commande_source": "", "multiple_commande_valeur": "",
        "minimum_commande": A, "panachage": A,
        # logistique
        "poids_brut_g": A, "longueur_cm": A, "largeur_cm": A, "hauteur_cm": A,
        "volume_dm3": A, "fragile": A, "contient_verre": A, "liquide": A,
        "sensible_chocs": A, "protection_necessaire": A, "protection_type": A,
        "contraintes_transport": A,
        # personnalisation
        "personnalisable": A, "type_personnalisation": A,
        "cout_personnalisation": A, "delai_personnalisation": A,
        "fournisseur_personnalisation": A,
        # commercial
        "niveau_gamme": "", "niveau_noblesse_propose": "",
        "noblesse_justification": "", "noblesse_statut": "",
        "saisonnalite": A, "produit_noel": "", "produit_premium": "",
        "produit_local_emblematique": "", "tags_opportunite": "",
        # stock
        "stock_interne": A, "stock_reserve": A, "stock_disponible": A,
        "stock_fournisseur": A, "delai_fournisseur": A,
        "quantite_minimum_commande": A, "disponibilite": "",
        # qualite des donnees
        "statut_ligne": "PRODUIT", "doublon_potentiel": "", "conflit_prix": "",
        "anomalie_source": "", "champs_a_confirmer": "",
        "source": "", "ligne_source": "",
        "_texte": "",
    }
    p.update(kw)
    return p


def finaliser(p, fournisseur_ref_pour_origine):
    """Derivations communes : origine, alcool, poids, logistique, noblesse."""
    texte = p["_texte"]
    origine = R.origine_info(texte, p["categorie"], p["sous_categorie"],
                             fournisseur_ref_pour_origine)
    origine["origine_departement"] = normaliser_departement(origine["origine_departement"])
    p.update(origine)

    alcool, degre = R.alcool_info(texte, p["categorie"], p["degre_alcool"])
    p["alcool"] = alcool
    p["degre_alcool"] = degre or p["degre_alcool"] or ""
    p["type_alcool"] = p["sous_categorie"] if alcool == "OUI" else (
        "Alcool cité comme ingrédient dans le libellé" if alcool == "INGREDIENT_ALCOOLISE" else "")

    if not p["poids_net_g"] and not p["volume_ml"]:
        poids, volume, mention = R.poids_volume(texte)
        p["poids_net_g"], p["volume_ml"], p["mention_poids_volume_source"] = poids, volume, mention

    p["alerte_matiere_premiere"] = R.alerte_matiere_premiere(texte, p["categorie"])
    p.update(R.logistique(texte, p["categorie"], p["volume_ml"]))
    PRICING.enrichir(p)

    niveau, justif = R.noblesse(p["prix_achat_ht"], p["categorie"], p["appellation"], texte)
    p["niveau_noblesse_propose"] = niveau
    p["noblesse_justification"] = justif
    p["noblesse_statut"] = "PROPOSITION AUTOMATIQUE – À VALIDER EN BACK-OFFICE"
    p["niveau_gamme"] = {1: "ESSENTIEL", 2: "ESSENTIEL", 3: "SIGNATURE",
                         4: "PREMIUM", 5: "PRESTIGE"}.get(niveau, A)

    if p["prix_achat_ht"] != "" and p["tva_taux"] not in ("", A):
        p["prix_achat_ttc"] = round(p["prix_achat_ht"] * (1 + float(p["tva_taux"])), 4)

    p["produit_noel"] = "OUI" if R.NOEL.search(texte) else A
    p["produit_premium"] = "OUI" if niveau not in ("", None) and int(niveau) >= 4 else "NON"
    p["produit_local_emblematique"] = "OUI" if p["caractere_local"] == "NORMANDIE_CONFIRMEE" else A
    p["tags_opportunite"] = " ; ".join(R.tags_opportunite(p))
    p["champs_a_confirmer"] = sum(1 for k, v in p.items() if not k.startswith("_") and v == A)
    return p


# --- 2.1 Tarif Normand Direct ---------------------------------------------
MULTIPLE = re.compile(r"(\d+)")
for row in read_csv("01_normand_direct_tarif_2026.csv"):
    libelle_a = row["fournisseur_colonne_A"]
    produit_lib = row["produit"]
    nom_prod, cond_suffixe = R.clean_producer_label(libelle_a)
    prod_ref = get_fournisseur(
        nom_prod, "PRODUCTEUR RÉFÉRENCÉ CHEZ LE DISTRIBUTEUR",
        "Via NORMAND DIRECT TERROIR", SRC_NORMAND, libelle_source=libelle_a)
    if row["note_fournisseur"]:
        prod_ref["conditions_relevees_dans_le_tarif"].add(row["note_fournisseur"])

    prix = num(row["pu_ht"])
    texte = f"{produit_lib} {libelle_a}"
    cat, sous, regle = R.classify_avec_defaut(produit_lib, libelle_a)

    statut = "PRODUIT"
    if not prix and prix != 0:
        if produit_lib.lower().startswith("disponible également"):
            statut = "LIGNE_INFORMATIVE_NON_PRODUIT"

    unite = "unité"
    prix_kg = ""
    if row["multiple_de_commande"].strip().lower() == "kg" or "€/kg" in row["pu_ht"]:
        unite = "kilogramme"
        prix_kg = prix if prix != "" else ""
    if produit_lib.strip().lower().startswith("disponible également"):
        unite = row["pu_ht"]

    mult_src = row["multiple_de_commande"]
    mult_val = MULTIPLE.search(mult_src).group(1) if MULTIPLE.search(mult_src) else ""

    p = base_produit(
        id_produit=new_id(),
        nom_produit=produit_lib,
        libelle_source=produit_lib,
        marque=nom_prod,
        id_producteur=prod_ref["id_fournisseur"],
        producteur=nom_prod,
        producteur_libelle_source=libelle_a,
        id_fournisseur=NORMAND["id_fournisseur"],
        fournisseur=NORMAND["nom"],
        canal_de_commande=NORMAND["canal_de_commande"],
        categorie=cat, sous_categorie=sous, regle_de_classification=regle,
        conditionnement=cond_suffixe or A,
        dlc_ddm=row["dlc_ddm"] or A,
        prix_achat_ht=prix,
        prix_achat_texte_source=row["pu_ht"],
        prix_au_kg_ht=prix_kg,
        unite_de_prix=unite,
        libelle_colonne_prix_source="PU HT",
        tva_taux=row["tva"] or A,
        multiple_commande_source=mult_src,
        multiple_commande_valeur=mult_val,
        minimum_commande=A,
        panachage=row["note_fournisseur"] or A,
        mention_source="NOUVEAU" if "NOUVEAU" in produit_lib.upper() else "",
        disponibilite="SUR_COMMANDE (document = tarif et bon de commande 2026)",
        statut_ligne=statut,
        source=SRC_NORMAND,
        ligne_source=f"ligne {row['ligne_source']}",
        _texte=texte,
    )
    if "mini 50 paquets" in produit_lib.lower():
        p["minimum_commande"] = "50 paquets (mention « mini 50 paquets » dans le libellé)"
    if row["note_fournisseur"]:
        p["panachage"] = row["note_fournisseur"]
    if statut != "PRODUIT":
        anomalies.append({
            "id_produit": p["id_produit"], "source": p["source"],
            "ligne_source": p["ligne_source"], "type": "LIGNE_INFORMATIVE",
            "detail": f"« {produit_lib} » : ligne de commentaire du tarif, sans prix unitaire "
                      f"(valeur colonne PU HT : « {row['pu_ht']} »). Conservée, non comptée comme produit.",
        })
    produits.append(finaliser(p, ref_producteur(nom_prod)))

# --- 2.2 Catalogue Caramels d'Isigny --------------------------------------
MINI = re.compile(r"(?i)(vrac \d+ ?kg minimum|boite de [^/]+|bo(î|i)te de \d+ unit(é|e)s[^/]*)")
for row in read_csv("02_caramels_isigny_2026.csv"):
    article = row["article"]
    gamme, sous_gamme = row["gamme"], row["sous_gamme"]
    cat, sous, regle = R.classify_isigny(gamme, article)

    lg, lk = row["libelle_colonne_G"].lower(), row["libelle_colonne_K"].lower()
    g, k = num(row["prix_ht_colonne_G"]), num(row["prix_colonne_K"])
    if "unitaire" in lg:
        prix, prix_kg = g, (k if "kg" in lk else "")
        libelle_prix = row["libelle_colonne_G"]
    elif "kg" in lg:
        prix, prix_kg = (k if "unitaire" in lk else ""), g
        libelle_prix = row["libelle_colonne_K"] or row["libelle_colonne_G"]
    else:
        prix, prix_kg = k, ""
        libelle_prix = row["libelle_colonne_K"]
    if prix == "" and k != "":
        prix, libelle_prix = k, row["libelle_colonne_K"]

    minimum = A
    m = MINI.search(sous_gamme)
    if m:
        minimum = f"{m.group(0)} (mention du catalogue : « {sous_gamme} »)"

    p = base_produit(
        id_produit=new_id(),
        reference_fournisseur=row["reference"],
        nom_produit=re.sub(r"^\d+\.", "", article).strip(),
        libelle_source=article,
        description=f"Gamme : {gamme}" + (f" – {sous_gamme}" if sous_gamme else ""),
        marque="CARAMELS D'ISIGNY",
        id_producteur=ISIGNY["id_fournisseur"],
        producteur="CARAMELS D'ISIGNY",
        producteur_libelle_source=gamme,
        id_fournisseur=ISIGNY["id_fournisseur"],
        fournisseur=ISIGNY["nom"],
        canal_de_commande=ISIGNY["canal_de_commande"],
        categorie=cat, sous_categorie=sous, regle_de_classification=regle,
        conditionnement=row["poids"] or A,
        quantite_par_conditionnement=row["colisage"] or "",
        prix_achat_ht=prix,
        prix_achat_texte_source=row["prix_ht_colonne_G"] or row["prix_colonne_K"],
        prix_au_kg_ht=prix_kg,
        libelle_colonne_prix_source=libelle_prix,
        tva_taux=row["tva"] or A,
        minimum_commande=minimum,
        disponibilite="SUR_COMMANDE (document = catalogue tarifé 2026)",
        source=SRC_ISIGNY,
        ligne_source=f"ligne {row['ligne_source']} – {gamme}",
        _texte=f"{article} {gamme} {sous_gamme}",
    )
    if row["ean"]:
        p["reference_fournisseur"] = f"{row['reference']} (EAN {row['ean']})"
    poids, volume, mention = R.poids_volume(row["poids"])
    p["poids_net_g"], p["volume_ml"], p["mention_poids_volume_source"] = poids, volume, row["poids"]
    produits.append(finaliser(p, ref_producteur("CARAMELS D'ISIGNY")))

# --- 2.3 Tarif Domaine de Billy -------------------------------------------
for row in read_csv("03_domaine_de_billy_tarif_2026.csv"):
    section, article = row["section"], row["article"]
    cat, sous, regle = R.classify_billy(section, article)
    prix = num(row["prix_ht"]) or num(row["prix_droits_inclus"])
    libelle_prix = row["libelle_colonne_prix_source"]

    p = base_produit(
        id_produit=new_id(),
        nom_produit=article,
        libelle_source=article,
        description=row["description"] or A,
        marque="DOMAINE DE BILLY",
        id_producteur=BILLY["id_fournisseur"],
        producteur="SARL DOMAINE DE BILLY",
        producteur_libelle_source=f"section {section}",
        id_fournisseur=BILLY["id_fournisseur"],
        fournisseur=BILLY["nom"],
        canal_de_commande=BILLY["canal_de_commande"],
        categorie=cat, sous_categorie=sous, regle_de_classification=regle,
        conditionnement=row["volume"] or A,
        prix_achat_ht=prix,
        prix_achat_texte_source=row["prix_ht"] or row["prix_droits_inclus"],
        libelle_colonne_prix_source=libelle_prix,
        prix_hors_droits_ht=num(row["prix_hors_droits"]),
        droits_accises=num(row["droits"]),
        tva_taux=A,
        degre_alcool=row["degre_ou_vat"],
        mention_source=row["mention"],
        disponibilite="SUR_COMMANDE (document = tarif producteur 2026)",
        source=f"{SRC_BILLY} – page {row['page_source']}",
        ligne_source=f"page {row['page_source']} – section {section}",
        _texte=f"{article} {row['description']} {section}",
    )
    if section == "APERITIFS" and not row["prix_hors_droits"] and row["prix_droits_inclus"]:
        p["anomalie_source"] = ("Seule la colonne « Prix de vente (droits inclus) » est renseignée "
                                "dans le document ; prix hors droits et montant des droits absents")
    poids, volume, mention = R.poids_volume(row["volume"])
    p["poids_net_g"], p["volume_ml"], p["mention_poids_volume_source"] = poids, volume, row["volume"]
    produits.append(finaliser(p, ref_producteur("SARL DOMAINE DE BILLY")))

# --- 2.4 Facture Domaine de Brucan ----------------------------------------
for row in read_csv("04_facture_brucan_2600000025.csv"):
    designation = row["designation"]
    cat, sous, ordre = R.classify(designation)
    pu_ttc = num(row["pu_ttc"])
    taux = num(row["taux_tva"]) / 100
    pu_ht = round(pu_ttc / (1 + taux), 4)

    p = base_produit(
        id_produit=new_id(),
        reference_fournisseur=row["article"],
        nom_produit=designation,
        libelle_source=designation,
        marque="DOMAINE DE BRUCAN",
        id_producteur=BRUCAN["id_fournisseur"],
        producteur="DOMAINE DE BRUCAN",
        producteur_libelle_source="facture",
        id_fournisseur=BRUCAN["id_fournisseur"],
        fournisseur=BRUCAN["nom"],
        canal_de_commande=BRUCAN["canal_de_commande"],
        categorie=cat, sous_categorie=sous if sous else "VIN",
        regle_de_classification=f"règle {ordre}" if ordre else "désignation de la facture",
        conditionnement="bouteille 75 cl (mention « btlle » sur la facture)",
        prix_achat_ht=pu_ht,
        prix_achat_texte_source=f"{row['pu_ttc']} TTC (prix unitaire facturé)",
        libelle_colonne_prix_source="P.u. T.T.C",
        prix_achat_ttc=pu_ttc,
        tva_taux=taux,
        disponibilite="À CONFIRMER (document = facture, pas un tarif catalogue)",
        anomalie_source=("Prix issu d'une FACTURE et non d'un tarif : le prix unitaire HT "
                         f"({pu_ht:.4f} €) est calculé à partir du prix TTC facturé "
                         f"({row['pu_ttc']}) et du taux de TVA de la facture (20 %). "
                         "Une remise globale de -46,62 € HT figure au pied de facture et n'est "
                         "PAS répercutée sur ce prix unitaire."),
        source=SRC_BRUCAN,
        ligne_source=f"ligne article {row['article']}",
        _texte=designation,
    )
    p["type_alcool"] = ("VIN – couleur déduite du nom pour « Blanc » et « Rosé » ; "
                        "cépage et appellation À CONFIRMER (code APE du domaine : 0121Z, culture de la vigne)")
    p["volume_ml"] = 750.0
    p["mention_poids_volume_source"] = "75 cl (désignation de la facture)"
    produits.append(finaliser(p, ref_producteur("DOMAINE DE BRUCAN")))

print(f"{len(produits)} lignes produits construites")


# ==========================================================================
# 3. Doublons, homonymies et conflits de prix
# ==========================================================================
def cle_produit(p):
    """Cle stricte : nom + format + gamme du catalogue. La gamme est
    determinante chez Caramels d'Isigny, ou le meme libelle existe en
    cellophane imprimee et en cellophane neutre a des prix differents."""
    return R.normalize_key(
        f"{p['nom_produit']} {p['conditionnement']} {p['mention_poids_volume_source']} "
        f"{p['description']}")


# Retire uniquement les tokens de FORMAT (volume, poids, age, degre, colisage,
# numero de catalogue) : les nombres porteurs de sens sont conserves, sinon
# « ASSORTIMENT 9 FRUITS » et « ASSORTIMENT 5 FRUITS » seraient rapproches.
FORMAT_TOKENS = re.compile(
    r"(?i)^\s*\d+\s*\.|"
    r"\b\d+(?:[,.]\d+)?\s*(?:kg|g|gr|grs|cl|ml|l|litres?|ans?)\b|"
    r"\d+(?:[,.]\d+)?\s*[°%]|"
    r"/\s*\d+")


def cle_libelle(p):
    """Cle large : nom seul, formats et numeros de catalogue retires. Sert a
    reperer les libelles identiques sans jamais fusionner les produits."""
    return R.normalize_key(FORMAT_TOKENS.sub(" ", p["nom_produit"]))


doublons, conflits = [], []
par_producteur = defaultdict(list)
par_nom = defaultdict(list)
for p in produits:
    if p["statut_ligne"] != "PRODUIT":
        continue
    par_producteur[(p["id_producteur"], cle_produit(p))].append(p)
    par_nom[cle_produit(p)].append(p)

for (id_prod, cle), groupe in par_producteur.items():
    if len(groupe) < 2:
        continue
    prix = {p["prix_achat_ht"] for p in groupe}
    type_ = "CONFLIT_DE_PRIX" if len(prix) > 1 else "DOUBLON_POTENTIEL_MEME_PRODUCTEUR"
    for p in groupe:
        p["doublon_potentiel"] = "DOUBLON POTENTIEL"
        if type_ == "CONFLIT_DE_PRIX":
            p["conflit_prix"] = "CONFLIT DE PRIX"
    ligne = {
        "type": type_,
        "cle_de_rapprochement": cle,
        "producteur": groupe[0]["producteur"],
        "ids_produits": " / ".join(p["id_produit"] for p in groupe),
        "libelles": " || ".join(p["libelle_source"] for p in groupe),
        "prix_ht": " / ".join(str(p["prix_achat_ht"]) for p in groupe),
        "sources": " || ".join(f"{p['source']} ({p['ligne_source']})" for p in groupe),
        "decision": "AUCUNE FUSION EFFECTUÉE – à arbitrer manuellement",
    }
    (conflits if type_ == "CONFLIT_DE_PRIX" else doublons).append(ligne)

par_libelle = defaultdict(list)
for p in produits:
    if p["statut_ligne"] == "PRODUIT" and cle_libelle(p):
        par_libelle[cle_libelle(p)].append(p)

for cle, groupe in sorted(par_libelle.items()):
    if len(groupe) < 2:
        continue
    deja = {tuple(sorted(p["id_produit"] for p in groupe))}
    if any(tuple(sorted(d["ids_produits"].split(" / "))) in deja for d in doublons + conflits):
        continue
    producteurs = sorted({p["producteur"] for p in groupe})
    memes = len(producteurs) == 1
    for p in groupe:
        if not p["doublon_potentiel"]:
            p["doublon_potentiel"] = ("LIBELLÉ IDENTIQUE – MÊME PRODUCTEUR" if memes
                                      else "LIBELLÉ IDENTIQUE – PRODUCTEURS DIFFÉRENTS")
    doublons.append({
        "type": "LIBELLE_IDENTIQUE_MEME_PRODUCTEUR" if memes
                else "LIBELLE_IDENTIQUE_PRODUCTEURS_DIFFERENTS",
        "cle_de_rapprochement": cle,
        "producteur": " / ".join(producteurs),
        "ids_produits": " / ".join(p["id_produit"] for p in groupe),
        "libelles": " || ".join(p["libelle_source"] for p in groupe),
        "prix_ht": " / ".join(str(p["prix_achat_ht"]) for p in groupe),
        "sources": " || ".join(f"{p['source']} ({p['ligne_source']})" for p in groupe),
        "decision": ("Même libellé, formats ou décors différents : NE PAS FUSIONNER sans vérification"
                     if memes else
                     "Produits de producteurs différents portant le même libellé : NE PAS FUSIONNER"),
    })

for cle, groupe in par_nom.items():
    producteurs = {p["id_producteur"] for p in groupe}
    if len(groupe) < 2 or len(producteurs) < 2:
        continue
    for p in groupe:
        if not p["doublon_potentiel"]:
            p["doublon_potentiel"] = "HOMONYMIE INTER-PRODUCTEURS"
    doublons.append({
        "type": "HOMONYMIE_INTER_PRODUCTEURS",
        "cle_de_rapprochement": cle,
        "producteur": " / ".join(sorted({p["producteur"] for p in groupe})),
        "ids_produits": " / ".join(p["id_produit"] for p in groupe),
        "libelles": " || ".join(p["libelle_source"] for p in groupe),
        "prix_ht": " / ".join(str(p["prix_achat_ht"]) for p in groupe),
        "sources": " || ".join(f"{p['source']} ({p['ligne_source']})" for p in groupe),
        "decision": "Produits de producteurs différents portant le même libellé : "
                    "NE PAS FUSIONNER, vérifier s'il s'agit du même article",
    })

# EAN identiques dans le catalogue Isigny (anomalie du document source)
ean_index = defaultdict(list)
for p in produits:
    m = re.search(r"EAN (\d+)", p["reference_fournisseur"] or "")
    if m:
        ean_index[m.group(1)].append(p)
for ean, groupe in ean_index.items():
    if len(groupe) > 1:
        for p in groupe:
            p["anomalie_source"] = (p["anomalie_source"] + " ; " if p["anomalie_source"] else "") + \
                f"EAN {ean} identique à un autre article du catalogue"
        anomalies.append({
            "id_produit": " / ".join(p["id_produit"] for p in groupe),
            "source": groupe[0]["source"],
            "ligne_source": " / ".join(p["ligne_source"] for p in groupe),
            "type": "EAN_DUPLIQUE_DANS_LE_DOCUMENT",
            "detail": f"L'EAN {ean} est affecté à {len(groupe)} articles différents : " +
                      " || ".join(f"{p['reference_fournisseur']} – {p['libelle_source']}" for p in groupe),
        })

# Incoherences de description relevees dans le tarif Domaine de Billy
for p in produits:
    if p["source"].startswith(SRC_BILLY) and "menthe" in p["nom_produit"].lower() \
            and "basilic" in (p["description"] or "").lower():
        p["anomalie_source"] = ("Description du document incohérente avec le nom du produit "
                                "(« Basilic frais infusé » pour un produit Menthe) – reprise telle quelle")
        anomalies.append({
            "id_produit": p["id_produit"], "source": p["source"],
            "ligne_source": p["ligne_source"], "type": "INCOHERENCE_DANS_LE_DOCUMENT_SOURCE",
            "detail": f"« {p['nom_produit']} » est décrit comme « {p['description']} ». "
                      "Donnée conservée telle quelle, à faire corriger par le fournisseur.",
        })


# ==========================================================================
# 4. Tarifs (une ligne par prix releve, avec sa source exacte)
# ==========================================================================
tarifs = []
for p in produits:
    if p["prix_achat_ht"] == "" and p["prix_au_kg_ht"] == "":
        continue
    tarifs.append({
        "id_tarif": f"TAR-{len(tarifs)+1:04d}",
        "id_produit": p["id_produit"],
        "nom_produit": p["nom_produit"],
        "id_fournisseur": p["id_fournisseur"],
        "fournisseur": p["fournisseur"],
        "prix_achat_ht": p["prix_achat_ht"],
        "unite_de_prix": p["unite_de_prix"],
        "prix_au_kg_ht": p["prix_au_kg_ht"],
        "prix_hors_droits_ht": p["prix_hors_droits_ht"],
        "droits_accises": p["droits_accises"],
        "tva_taux": p["tva_taux"],
        "prix_achat_ttc": p["prix_achat_ttc"],
        "texte_exact_du_document": p["prix_achat_texte_source"],
        "libelle_colonne_prix_source": p["libelle_colonne_prix_source"],
        "validite": "Tarif 2026" if "2026" in p["source"] else A,
        "type_de_document": "FACTURE (prix réellement facturé)" if "facture" in p["source"].lower()
                            else "TARIF FOURNISSEUR",
        "source": p["source"],
        "ligne_source": p["ligne_source"],
    })

# ==========================================================================
# 5. Remises et avantages commerciaux (uniquement ceux presents dans les documents)
# ==========================================================================
remises = [
    {
        "id_remise": "REM-0001", "type": "PALIER_QUANTITATIF",
        "fournisseur": "NORMAND DIRECT TERROIR", "producteur": "BISCUITERIE DE QUINEVILLE",
        "produits_concernes": "Boutons de la Manche 250 grs",
        "palier": "moins de 12 unités", "prix_ht": "3.60", "remise": "",
        "texte_exact_du_document": "Boutons de la Manche 250 grs moins de 12 – PU HT 3,60 – multiple 6/",
        "source": SRC_NORMAND, "ligne_source": "ligne 14",
    },
    {
        "id_remise": "REM-0002", "type": "PALIER_QUANTITATIF",
        "fournisseur": "NORMAND DIRECT TERROIR", "producteur": "BISCUITERIE DE QUINEVILLE",
        "produits_concernes": "Boutons de la Manche 250 grs",
        "palier": "carton de 12", "prix_ht": "3.45",
        "remise": "-0,15 € HT par unité soit -4,17 % par rapport au palier « moins de 12 »",
        "texte_exact_du_document": "Boutons de la Manche 250 grs/ carton de 12 – PU HT 3,45 – multiple 12/",
        "source": SRC_NORMAND, "ligne_source": "ligne 15",
    },
    {
        "id_remise": "REM-0003", "type": "DEDUCTION_SUR_FACTURE",
        "fournisseur": "DOMAINE DE BRUCAN", "producteur": "DOMAINE DE BRUCAN",
        "produits_concernes": "Ensemble de la facture n° 2600000025",
        "palier": "Commande de 42 bouteilles – 466,20 € HT",
        "prix_ht": "", "remise": "-46,62 € HT (soit exactement 10,00 % du total HT)",
        "texte_exact_du_document": "Total H.T. 466,20 € / ligne « Montant H.T. » -46,62 € / Net H.T. 419,58 €",
        "source": SRC_BRUCAN,
        "ligne_source": "pied de facture – LE LIBELLÉ DE LA DÉDUCTION N'EST PAS EXPLICITE DANS LE "
                        "DOCUMENT : s'agit-il d'une remise commerciale, d'une remise de fin d'année "
                        "ou d'un avoir ? À CONFIRMER auprès du fournisseur",
    },
    {
        "id_remise": "AVA-0001", "type": "AVANTAGE_COMMERCIAL (échantillon)",
        "fournisseur": "NORMAND DIRECT TERROIR", "producteur": "DISTILLERIE C'EST NOUS",
        "produits_concernes": "Gin 40° 70 cl / Gin Concombre-basilic-pomme / Vodka 40° 70 cl / "
                              "Liqueur de café Cold brew 70 cl / Limoncello 50 cl / Pastis du verger",
        "palier": "pour 6 bouteilles", "prix_ht": "",
        "remise": "Échantillon gratuit sur demande",
        "texte_exact_du_document": "« Echantillon gratuit pour 6 sur demande » (mention dans le libellé produit)",
        "source": SRC_NORMAND, "ligne_source": "lignes 413, 415, 418, 420, 422, 427",
    },
    {
        "id_remise": "AVA-0002", "type": "AVANTAGE_COMMERCIAL (emballage offert)",
        "fournisseur": "NORMAND DIRECT TERROIR", "producteur": "POM' POM'",
        "produits_concernes": "Coffret 2 x 20 cl",
        "palier": "avec l'achat de bouteilles 20 cl", "prix_ht": "0",
        "remise": "Coffret facturé 0,00 € HT",
        "texte_exact_du_document": "« Coffret 2 x 20cl INCLUS AVEC LES 20 CL SUR DEMANDE » – PU HT 0",
        "source": SRC_NORMAND, "ligne_source": "ligne 316",
    },
]

# ==========================================================================
# 6. Conditions commerciales relevees par producteur
# ==========================================================================
conditions = []
for f in fournisseurs.values():
    for c in sorted(f["conditions_relevees_dans_le_tarif"]):
        conditions.append({
            "id_fournisseur": f["id_fournisseur"], "fournisseur": f["nom"],
            "type": "PANACHAGE / EXPEDITION / MINIMUM",
            "texte_exact_du_document": c,
            "source": SRC_NORMAND,
        })

# ==========================================================================
# 7. Options / extras vendables en complement
# ==========================================================================
CATS_OPTION = {
    "COFFRET_ET_EMBALLAGE": "COFFRET_OU_EMBALLAGE",
    "ART_DE_LA_TABLE_ET_OBJET": "OBJET_ET_ART_DE_LA_TABLE",
    "COSMETIQUE_BIEN_ETRE": "PRODUIT_NON_ALIMENTAIRE",
}
options = []
for p in produits:
    if p["categorie"] not in CATS_OPTION or p["statut_ligne"] != "PRODUIT":
        continue
    options.append({
        "id_option": f"OPT-{len(options)+1:04d}",
        "id_produit": p["id_produit"],
        "nom": p["nom_produit"],
        "type_option": CATS_OPTION[p["categorie"]],
        "fournisseur": p["fournisseur"],
        "producteur": p["producteur"],
        "prix_achat_ht": p["prix_achat_ht"],
        "tva_taux": p["tva_taux"],
        "prix_vente_ht": "À DÉFINIR",
        "remarque": "Nature exacte à confirmer : coffret garni (produit) ou emballage seul"
                    if p["categorie"] == "COFFRET_ET_EMBALLAGE" else "",
        "source": p["source"], "ligne_source": p["ligne_source"],
    })


# ==========================================================================
# 8. Scores et opportunites
# ==========================================================================
scores_rows, opportunites_rows = [], []
for p in produits:
    if p["statut_ligne"] != "PRODUIT":
        continue
    s = R.scores(p)
    scores_rows.append({
        "id_produit": p["id_produit"], "nom_produit": p["nom_produit"],
        **{k: v for k, v in s.items()},
        "score_marge_commentaire": R.NON_CALCULABLE + " : aucun prix de vente n'existe encore",
        "score_qualite_commentaire": R.NON_CALCULABLE + " : à saisir en back-office",
        "statut": "PROPOSITION AUTOMATIQUE – À VALIDER EN BACK-OFFICE",
        "base_de_calcul": p["noblesse_justification"],
    })
    for tag in R.tags_opportunite(p):
        opportunites_rows.append({
            "id_produit": p["id_produit"], "nom_produit": p["nom_produit"],
            "tag": tag, "origine_du_tag": "Calculé à partir des données documentées",
            "statut": "À VALIDER",
        })

# ==========================================================================
# 9. Alternatives produits
# ==========================================================================
CATS_SANS_ALTERNATIVE = {"INGREDIENT_PROFESSIONNEL", "COFFRET_ET_EMBALLAGE"}
alternatives = []
candidats = [p for p in produits
             if p["statut_ligne"] == "PRODUIT" and p["prix_achat_ht"] != ""
             and p["categorie"] not in CATS_SANS_ALTERNATIVE]
par_sous_cat = defaultdict(list)
for p in candidats:
    par_sous_cat[(p["sous_categorie"], p["alcool"])].append(p)

for p in candidats:
    prix = float(p["prix_achat_ht"])
    if prix <= 0:
        continue
    pool = []
    for q in par_sous_cat[(p["sous_categorie"], p["alcool"])]:
        if q["id_produit"] == p["id_produit"] or q["prix_achat_ht"] in ("", 0):
            continue
        pq = float(q["prix_achat_ht"])
        ecart = (pq - prix) / prix
        if abs(ecart) > 0.40:
            continue
        pool.append((abs(ecart), ecart, q))
    for _, ecart, q in sorted(pool, key=lambda x: x[0])[:5]:
        alternatives.append({
            "id_produit": p["id_produit"], "nom_produit": p["nom_produit"],
            "id_alternative": q["id_produit"], "nom_alternative": q["nom_produit"],
            "relation_prix": "PRIX_INFERIEUR" if ecart < -0.02 else
                             ("PRIX_SUPERIEUR" if ecart > 0.02 else "PRIX_EQUIVALENT"),
            "ecart_prix_pct": round(ecart * 100, 1),
            "prix_produit_ht": p["prix_achat_ht"], "prix_alternative_ht": q["prix_achat_ht"],
            "meme_sous_categorie": "OUI", "meme_regime_alcool": "OUI",
            "meme_producteur": "OUI" if p["id_producteur"] == q["id_producteur"] else "NON",
            "meme_region_documentee": "OUI" if p["caractere_local"] == q["caractere_local"] else "NON",
            "ecart_noblesse": (int(q["niveau_noblesse_propose"]) - int(p["niveau_noblesse_propose"]))
            if p["niveau_noblesse_propose"] not in ("", None) and q["niveau_noblesse_propose"] not in ("", None) else "",
            "critere_de_rapprochement": "même sous-catégorie + même régime alcool + écart de prix ≤ 40 %",
        })

# ==========================================================================
# 10. Coffrets
# ==========================================================================
coffrets, coffrets_lignes = [], []
for p in produits:
    if p["categorie"] != "COFFRET_ET_EMBALLAGE":
        continue
    coffrets.append({
        "id_coffret": f"COF-{len(coffrets)+1:04d}",
        "nom": p["nom_produit"], "type": "COFFRET FOURNISSEUR (prêt à vendre)",
        "gamme": A, "id_fournisseur": p["id_fournisseur"], "fournisseur": p["fournisseur"],
        "producteur": p["producteur"],
        "composition_documentee": p["libelle_source"],
        "cout_produits_ht": p["prix_achat_ht"], "cout_emballage_ht": "À DÉFINIR",
        "cout_protection_ht": "À DÉFINIR", "cout_calage_ht": "À DÉFINIR",
        "cout_personnalisation_ht": "À DÉFINIR", "cout_preparation_ht": "À DÉFINIR",
        "cout_transport_ht": "À DÉFINIR", "cout_revient_total_ht": "À DÉFINIR",
        "prix_vente_ht": "À DÉFINIR", "remise": "À DÉFINIR", "prix_final_ht": "À DÉFINIR",
        "marge_eur": "À DÉFINIR", "taux_marge": "À DÉFINIR",
        "poids_total_g": p["poids_net_g"] or A, "dimensions": A,
        "source": p["source"], "ligne_source": p["ligne_source"],
        "remarque": "Nature à confirmer (coffret garni ou emballage seul). "
                    "Contenu détaillé non précisé dans le document.",
    })
    coffrets_lignes.append({
        "id_coffret": coffrets[-1]["id_coffret"], "id_produit": p["id_produit"],
        "nom_produit": p["nom_produit"], "quantite": 1,
        "role_dans_le_coffret": "COFFRET COMPLET FOURNISSEUR",
        "source": p["source"],
    })

for gamme, libelle in [("ESSENTIEL", "OPTION 1 – entrée de gamme"),
                       ("SIGNATURE", "OPTION 2 – meilleure valeur / recommandée"),
                       ("PRESTIGE", "OPTION 3 – premium")]:
    coffrets.append({
        "id_coffret": f"COF-{len(coffrets)+1:04d}",
        "nom": f"MODÈLE {gamme}", "type": "GABARIT INTERNE (à composer)",
        "gamme": libelle, "id_fournisseur": "", "fournisseur": "Composition interne",
        "producteur": "", "composition_documentee": "VIDE – à composer par le moteur",
        "cout_produits_ht": "À CALCULER", "cout_emballage_ht": "À CALCULER",
        "cout_protection_ht": "À CALCULER", "cout_calage_ht": "À CALCULER",
        "cout_personnalisation_ht": "À CALCULER", "cout_preparation_ht": "À CALCULER",
        "cout_transport_ht": "À CALCULER", "cout_revient_total_ht": "À CALCULER",
        "prix_vente_ht": "À DÉFINIR", "remise": "À DÉFINIR", "prix_final_ht": "À CALCULER",
        "marge_eur": "À CALCULER", "taux_marge": "À CALCULER",
        "poids_total_g": "À CALCULER", "dimensions": "À CALCULER",
        "source": "Structure demandée par le client (logique 3 niveaux)",
        "ligne_source": "",
        "remarque": "Aucune composition n'a été inventée : le gabarit est vide tant que les "
                    "paramètres de coût (emballage, préparation, transport) et les prix de "
                    "vente ne sont pas renseignés.",
    })

# ==========================================================================
# 11. Tables a alimenter (aucune donnee dans les documents fournis)
# ==========================================================================
emballages = [{
    "id_emballage": "", "reference": "", "id_fournisseur_emballage": "", "nom": "", "type": "",
    "longueur_cm": "", "largeur_cm": "", "hauteur_cm": "", "volume_utile_dm3": "", "poids_g": "",
    "prix_achat_ht": "", "quantite_par_conditionnement": "", "prix_unitaire_ht": "",
    "delai": "", "stock": "", "minimum_commande": "", "compatible_avec": "", "source": "",
}]
emballages_types = [
    "coffret carton", "boîte rigide", "caisse", "panier", "boîte cadeau", "étui", "sachet",
    "papier de soie", "frisure", "calage", "ruban", "carte", "étiquette", "autocollant",
    "protection bouteille", "protection verre", "papier bulle / protection adaptée",
    "séparateur", "intercalaire", "enveloppe", "emballage extérieur d'expédition",
]
emballages_fournisseurs = [
    {"id_fournisseur_emballage": "EMB-FOUR-0001", "nom": "RETIF", "site_internet": "À RENSEIGNER",
     "contact": "À RENSEIGNER", "conditions": "À RENSEIGNER", "minimum_commande": "À RENSEIGNER",
     "delai": "À RENSEIGNER", "source": "Piste citée par le client – aucun tarif fourni"},
    {"id_fournisseur_emballage": "EMB-FOUR-0002", "nom": "RAJA", "site_internet": "À RENSEIGNER",
     "contact": "À RENSEIGNER", "conditions": "À RENSEIGNER", "minimum_commande": "À RENSEIGNER",
     "delai": "À RENSEIGNER", "source": "Piste citée par le client – aucun tarif fourni"},
]
protections = [{
    "id_protection": "", "nom": "", "type": "", "declencheur": "", "cout_unitaire_ht": "",
    "poids_g": "", "id_emballage_associe": "", "source": "",
}]
transport = [{
    "id_tarif_transport": "", "transporteur": "", "service": "", "zone": "",
    "poids_min_kg": "", "poids_max_kg": "", "tarif_ht": "", "supplement": "",
    "dimensions_max": "", "poids_volumetrique": "", "assurance": "", "delai": "", "source": "",
}]
personnalisation = [{
    "id_personnalisation": "PERS-0001",
    "type": "AUCUNE DONNÉE DANS LES DOCUMENTS FOURNIS",
    "produits_concernes": "—",
    "personnalisation_rapide": A, "personnalisation_avec_delai_fournisseur": A,
    "delai": A, "cout_ht": A, "fournisseur_de_personnalisation": A,
    "commentaire": "Aucun des quatre documents ne mentionne de marquage, gravure, étiquette "
                   "personnalisée ni logo client. Les seules souplesses documentées sont des "
                   "règles de PANACHAGE et de choix de parfums à la commande, listées dans "
                   "l'onglet CONDITIONS_FOURNISSEURS.",
    "source": "Constat d'absence – à vérifier auprès de chaque fournisseur",
}]

parametres = [
    ("COUT_PREPARATION_COFFRET_SIMPLE", "À DÉFINIR", "minutes", "Temps de préparation d'un coffret simple"),
    ("COUT_PREPARATION_COFFRET_COMPLEXE", "À DÉFINIR", "minutes", "Temps de préparation d'un coffret complexe"),
    ("COUT_PREPARATION_COFFRET_PERSONNALISE", "À DÉFINIR", "minutes", "Temps de préparation d'un coffret personnalisé"),
    ("TAUX_HORAIRE_PREPARATION", "À DÉFINIR", "€ HT / heure", "Coût horaire de main-d'œuvre"),
    ("TAUX_MARGE_CIBLE_ESSENTIEL", "À DÉFINIR", "%", "Taux de marge visé sur la gamme Essentiel"),
    ("TAUX_MARGE_CIBLE_SIGNATURE", "À DÉFINIR", "%", "Taux de marge visé sur la gamme Signature"),
    ("TAUX_MARGE_CIBLE_PRESTIGE", "À DÉFINIR", "%", "Taux de marge visé sur la gamme Prestige"),
    ("TOLERANCE_BUDGET_INFERIEURE", "À DÉFINIR", "%", "Écart accepté sous le budget client"),
    ("TOLERANCE_BUDGET_SUPERIEURE", "À DÉFINIR", "%", "Écart accepté au-dessus du budget client"),
    ("DIVISEUR_POIDS_VOLUMETRIQUE", "À DÉFINIR", "—", "Diviseur transporteur pour le poids volumétrique"),
    ("SEUIL_PRODUIT_LOURD", "1000", "g", "Seuil utilisé pour le tag PRODUIT_LOURD (paramétrable)"),
    ("SEUIL_FAIBLE_COUT_LOGISTIQUE", "300", "g", "Seuil utilisé pour le tag FAIBLE_COUT_LOGISTIQUE"),
    ("PALIERS_NOBLESSE_PRIX", "2,50 / 5 / 10 / 25 €", "€ HT", "Paliers de prix d'achat utilisés pour la noblesse proposée"),
    ("ECART_MAX_ALTERNATIVE", "40", "%", "Écart de prix maximal entre un produit et son alternative"),
    ("PRIORITE_MOTEUR", "1 besoin client, 2 budget, 3 cohérence, 4 qualité perçue, "
                        "5 niveau de gamme, 6 disponibilité, 7 logistique, 8 marge", "—",
     "Hiérarchie de décision demandée par le client"),
]
parametres_rows = [{"parametre": k, "valeur": v, "unite": u, "description": d,
                    "modifiable_en_back_office": "OUI"} for k, v, u, d in parametres]


# ==========================================================================
# 12. Controle des donnees
# ==========================================================================
vrais_produits = [p for p in produits if p["statut_ligne"] == "PRODUIT"]


def _mediane(valeurs):
    if not valeurs:
        return 0.0
    tri = sorted(valeurs)
    milieu = len(tri) // 2
    return tri[milieu] if len(tri) % 2 else (tri[milieu - 1] + tri[milieu]) / 2


def compte(pred):
    return sum(1 for p in vrais_produits if pred(p))


controle = [
    ("Nombre de documents sources analysés", 3,
     "1 classeur Excel (2 onglets), 2 PDF – voir onglet SOURCES"),
    ("Nombre de fournisseurs / producteurs identifiés", len(fournisseurs),
     "Inclut 2 distributeurs/fabricants et les producteurs listés en colonne A du tarif Normand Direct"),
    ("dont canaux de commande directs identifiés", 4,
     "Normand Direct Terroir, Caramels d'Isigny (canal à confirmer), Domaine de Billy, Domaine de Brucan"),
    ("Nombre de produits", len(vrais_produits), "Lignes de statut PRODUIT"),
    ("Lignes informatives non comptées comme produits",
     len(produits) - len(vrais_produits), "Conservées dans l'onglet PRODUITS avec leur statut"),
    ("Nombre de références tarifées", len(tarifs), "Lignes de l'onglet TARIFS"),
    ("Produits sans prix", compte(lambda p: p["prix_achat_ht"] == "" and p["prix_au_kg_ht"] == ""), ""),
    ("Produits sans fournisseur", compte(lambda p: not p["id_fournisseur"]), ""),
    ("Produits sans producteur identifié", compte(lambda p: p["producteur"] == A), ""),
    ("Produits sans TVA indiquée", compte(lambda p: p["tva_taux"] == A), ""),
    ("Produits sans origine documentée",
     compte(lambda p: p["niveau_preuve_origine"] in
            ("AUCUNE_INFORMATION_DANS_LES_DOCUMENTS", "INDICE_DANS_LE_NOM_DU_PRODUIT_SEULEMENT")),
     "Règle stricte : un produit n'est pas normand parce qu'il est distribué par une entreprise normande"),
    ("Produits normands confirmés",
     compte(lambda p: p["caractere_local"] == "NORMANDIE_CONFIRMEE"),
     "Preuve = adresse du producteur dans son propre tarif, ou appellation protégée citée"),
    ("dont Calvados (14)", compte(lambda p: "Calvados" in str(p["origine_departement"])), ""),
    ("dont Manche (50)", compte(lambda p: "Manche" in str(p["origine_departement"])), ""),
    ("dont Orne (61)", compte(lambda p: "Orne" in str(p["origine_departement"])), ""),
    ("dont Eure (27)", compte(lambda p: "Eure" in str(p["origine_departement"])), ""),
    ("dont Seine-Maritime (76)",
     compte(lambda p: "Seine-Maritime" in str(p["origine_departement"])), ""),
    ("Produits transformés en Normandie à partir d'une matière première non française",
     compte(lambda p: bool(p["alerte_matiere_premiere"])),
     "Café, thé, cacao, fruits exotiques : à ne pas présenter comme « 100 % local »"),
    ("dont département normand à confirmer",
     compte(lambda p: p["caractere_local"] == "NORMANDIE_CONFIRMEE" and p["origine_departement"] == A), ""),
    ("Produits portant un indice de normandité dans leur seul nom",
     compte(lambda p: p["caractere_local"] == "NORMANDIE_INDIQUEE_DANS_LE_NOM_A_VERIFIER"),
     "À vérifier : un nom commercial n'est pas une preuve d'origine"),
    ("Produits français confirmés", compte(lambda p: p["origine_pays"] == "France"), ""),
    ("Produits d'origine non déterminée", compte(lambda p: p["origine_pays"] == A), ""),
    ("Produits AVEC alcool (boisson alcoolisée)", compte(lambda p: p["alcool"] == "OUI"), ""),
    ("Produits SANS alcool", compte(lambda p: p["alcool"] == "NON"), ""),
    ("Produits contenant de l'alcool comme ingrédient",
     compte(lambda p: p["alcool"] == "INGREDIENT_ALCOOLISE"),
     "Ni « avec alcool » ni totalement neutre : à arbitrer pour les coffrets sans alcool"),
    ("Produits personnalisables (documenté)", compte(lambda p: p["personnalisable"] == "OUI"),
     "Aucun document fourni ne mentionne de personnalisation (marquage, gravure, étiquette)"),
    ("Produits dont la personnalisation reste à confirmer",
     compte(lambda p: p["personnalisable"] == A), ""),
    ("Règles de panachage / conditions relevées", len(conditions), "Onglet CONDITIONS_FOURNISSEURS"),
    ("Doublons potentiels (lignes de rapprochement)", len(doublons),
     "Aucune fusion effectuée – voir onglet DOUBLONS"),
    ("dont homonymies entre producteurs différents",
     sum(1 for d in doublons if d["type"] == "HOMONYMIE_INTER_PRODUCTEURS"), ""),
    ("Conflits de prix", len(conflits), "Voir onglet CONFLITS"),
    ("Anomalies relevées dans les documents sources", len(anomalies), "Voir onglet ANOMALIES"),
    ("Règles de remise documentées",
     sum(1 for r in remises if r["type"].startswith("PALIER") or r["type"].startswith("DEDUCTION")),
     "Aucune grille de remise quantitative (1-49 / 50-99 / 100-199…) n'existe dans les documents fournis"),
    ("Avantages commerciaux documentés",
     sum(1 for r in remises if r["type"].startswith("AVANTAGE")), ""),
    ("Produits avec poids net lu dans le document",
     compte(lambda p: p["poids_net_g"] != ""), ""),
    ("Produits avec volume lu dans le document",
     compte(lambda p: p["volume_ml"] != ""), ""),
    ("Produits sans poids ni volume", compte(lambda p: not p["poids_net_g"] and not p["volume_ml"]), ""),
    ("Produits avec dimensions renseignées", 0,
     "Aucun document ne contient de dimensions : longueur/largeur/hauteur à mesurer ou à demander"),
    ("Produits avec prix de vente proposé", compte(lambda p: p["prix_vente_ht"] != ""),
     "PROPOSITION : coefficients de marché (build/coefficients_prix.csv), non issus des documents"),
    ("Taux de marge médian sur le prix de vente proposé",
     f"{_mediane([float(p['taux_marge']) for p in vrais_produits if p['taux_marge']]) * 100:.1f} %",
     "À valider avant toute mise en vente"),
    ("Produits sans prix de vente", compte(lambda p: p["prix_vente_ht"] == ""),
     "Produits dont le prix d'achat est nul ou absent"),
    ("Emballages référencés", 0, "Base EMBALLAGES à alimenter (aucun tarif emballage fourni)"),
    ("Tarifs de transport référencés", 0, "Base TARIFS_TRANSPORT à alimenter"),
    ("Alternatives produits calculées", len(alternatives),
     "Même sous-catégorie, même régime alcool, écart de prix ≤ 40 %"),
    ("Total de champs marqués À CONFIRMER",
     sum(int(p["champs_a_confirmer"]) for p in vrais_produits), ""),
]
controle_rows = [{"indicateur": k, "valeur": v, "commentaire": c} for k, v, c in controle]

repartition = Counter((p["categorie"], p["sous_categorie"]) for p in vrais_produits)
repartition_rows = [{"categorie": c, "sous_categorie": s, "nombre_de_produits": n}
                    for (c, s), n in sorted(repartition.items(), key=lambda x: (-x[1], x[0]))]

sources_rows = [
    {"id_source": "SRC-01", "document": "Tarif_et_bon_de_commande_2026.xlsx",
     "emetteur": "NORMAND DIRECT TERROIR", "nature": "Tarif professionnel + bon de commande",
     "perimetre": "Onglet « TARIF NORMAND DIRECT 2026 » – lignes 10 à 512",
     "date_ou_version": "2026 (fichier modifié le 24/04/2026, dernier auteur : Sylvain Le cornec)",
     "elements_repris": "Producteur, DLC/DDM, libellé produit, PU HT, multiple de commande, TVA, notes de panachage"},
    {"id_source": "SRC-02", "document": "Tarif_et_bon_de_commande_2026.xlsx",
     "emetteur": "CARAMELS D'ISIGNY", "nature": "Catalogue tarifé + bon de commande",
     "perimetre": "Onglet « CARAMELS ISIGNY 2026 » – renvois aux pages 4 à 21 du catalogue papier (non fourni)",
     "date_ou_version": "2026",
     "elements_repris": "Gamme, article, référence NC, poids, prix HT, TVA, EAN, colisage"},
    {"id_source": "SRC-03", "document": "billy.pdf",
     "emetteur": "SARL DOMAINE DE BILLY", "nature": "Conditions générales de vente + tarif producteur",
     "perimetre": "Pages 1 à 3 : cidres, poiré, vinaigre, jus, apéritifs, calvados, coffrets",
     "date_ou_version": "« Prix de vente à partir de 2026 »",
     "elements_repris": "Coordonnées complètes du producteur, produits, degrés, volumes, descriptions, "
                        "prix HT, droits d'accises. Coordonnées bancaires volontairement non reprises."},
    {"id_source": "SRC-04", "document": "FACT_2600000025_THE_ART_OF_WINE_ACADEMY.pdf",
     "emetteur": "DOMAINE DE BRUCAN", "nature": "Facture (et non tarif)",
     "perimetre": "Page 1 – 4 références de vin, pied de facture",
     "date_ou_version": "Facture n° 2600000025 du 12/08/2026",
     "elements_repris": "Coordonnées et identifiants du domaine, 4 articles avec prix unitaire TTC, "
                        "déduction globale de -46,62 € HT, conditions de règlement. "
                        "Coordonnées bancaires volontairement non reprises."},
]

# ==========================================================================
# 13. Export CSV + Excel
# ==========================================================================
def export(nom, rows, colonnes=None):
    EXPORT.mkdir(parents=True, exist_ok=True)
    if not rows:
        rows = []
    cols = colonnes or (list(rows[0].keys()) if rows else [])
    cols = [c for c in cols if not c.startswith("_")]
    with open(EXPORT / f"{nom}.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    return cols


fournisseurs_rows = []
for f in sorted(fournisseurs.values(), key=lambda x: x["id_fournisseur"]):
    row = dict(f)
    row["libelles_source"] = " | ".join(f["libelles_source"])
    row["conditions_relevees_dans_le_tarif"] = " | ".join(sorted(f["conditions_relevees_dans_le_tarif"]))
    row["nombre_de_produits"] = sum(
        1 for p in vrais_produits if p["id_producteur"] == f["id_fournisseur"])
    fournisseurs_rows.append(row)

TABLES = [
    ("PRODUITS", produits),
    ("FOURNISSEURS", fournisseurs_rows),
    ("TARIFS", tarifs),
    ("REMISES", remises),
    ("OPTIONS", options),
    ("COFFRETS", coffrets),
    ("COFFRETS_LIGNES", coffrets_lignes),
    ("CONTROLE_DES_DONNEES", controle_rows),
    ("REPARTITION_CATEGORIES", repartition_rows),
    ("PRODUITS_SCORES", scores_rows),
    ("PRODUITS_OPPORTUNITES", opportunites_rows),
    ("PRODUITS_ALTERNATIVES", alternatives),
    ("CONDITIONS_FOURNISSEURS", conditions),
    ("PERSONNALISATION", personnalisation),
    ("DOUBLONS", doublons),
    ("CONFLITS", conflits),
    ("ANOMALIES", anomalies),
    ("EMBALLAGES", emballages),
    ("EMBALLAGES_FOURNISSEURS", emballages_fournisseurs),
    ("EMBALLAGES_TYPES", [{"type_emballage": t, "statut": "À RENSEIGNER"} for t in emballages_types]),
    ("PROTECTIONS", protections),
    ("TARIFS_TRANSPORT", transport),
    ("PARAMETRES_MOTEUR", parametres_rows),
    ("SOURCES", sources_rows),
]

for nom, rows in TABLES:
    export(nom, rows)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    entete = Font(bold=True, color="FFFFFF")
    fond = PatternFill("solid", fgColor="1F4E5F")
    for nom, rows in TABLES:
        ws = wb.create_sheet(nom[:31])
        cols = [c for c in (list(rows[0].keys()) if rows else ["(table vide)"])
                if not c.startswith("_")]
        ws.append(cols)
        for cell in ws[1]:
            cell.font, cell.fill = entete, fond
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(cols, start=1):
            largeur = max(12, min(46, len(str(c)) + 4,
                                  max([len(str(c)) + 4] + [len(str(r.get(c, ""))) + 2
                                                           for r in rows[:200]])))
            ws.column_dimensions[get_column_letter(i)].width = largeur
    out = EXPORT / "BASE_PRODUITS_FOURNISSEURS_2026.xlsx"
    wb.save(out)
    print(f"Excel : {out}")
except ImportError:
    print("openpyxl absent : export Excel ignoré, les CSV sont générés")

print(f"{len(TABLES)} tables exportées dans {EXPORT}")
for k, v, c in controle[:6]:
    print(f"  {k} : {v}")
